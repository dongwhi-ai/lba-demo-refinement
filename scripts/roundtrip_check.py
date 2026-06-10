# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl"]
# ///
"""export된 xlsx를 원본과 전수 비교한다.

실행: uv run scripts/roundtrip_check.py <원본.xlsx> <export본.xlsx> <annotation.json>
검증: refinement(N)·검수상태(Q) 외 모든 셀의 값/타입 일치, N·Q는 검수 규칙과 일치.
"""

import json
import sys

import openpyxl

CATEGORY_ORDER = {"Aa": 0, "Ab": 1, "Ac": 2, "Ba": 3, "Bb": 4, "C": 5}
STATUS_LABELS = {"accept": "채택", "fix": "수정", "del": "폐기", None: "미검수"}
REFINEMENT_COL = 14  # N (1-based)
STATUS_COL = 17  # Q
QA_IDX_COL = 2  # B


def expected_refinement(ann: dict | None) -> str:
    if ann is None or ann["s"] == "accept":
        return ""
    if ann["s"] == "del":
        return "del"
    cats = sorted(ann.get("c") or [], key=CATEGORY_ORDER.__getitem__)
    memo = (ann.get("m") or "").strip()
    return ", ".join(["fix", *cats, *([memo] if memo else [])])


def cell_value(ws, row, col):
    return ws.cell(row=row, column=col).value


def main() -> None:
    orig_path, export_path, ann_path = sys.argv[1:4]
    orig = openpyxl.load_workbook(orig_path)[
        openpyxl.load_workbook(orig_path).sheetnames[0]
    ]
    exported_wb = openpyxl.load_workbook(export_path)
    assert exported_wb.sheetnames == ["sc04_results"], exported_wb.sheetnames
    exported = exported_wb["sc04_results"]
    ann_map = json.loads(open(ann_path, encoding="utf-8").read())

    assert orig.max_row == exported.max_row == 501, (orig.max_row, exported.max_row)

    diffs = []
    for r in range(1, 502):
        for c in range(1, 17):
            o, e = cell_value(orig, r, c), cell_value(exported, r, c)
            if c == REFINEMENT_COL and r > 1:
                qa = str(cell_value(orig, r, QA_IDX_COL))
                want = expected_refinement(ann_map.get(qa)) or None
                if e != want:
                    diffs.append((r, c, "refinement", want, e))
                continue
            if o != e or type(o) is not type(e):
                diffs.append((r, c, "mismatch", repr(o), repr(e)))
        # 검수상태 보조열
        qa = str(cell_value(orig, r, QA_IDX_COL))
        status_cell = cell_value(exported, r, STATUS_COL)
        if r == 1:
            if status_cell != "검수상태":
                diffs.append((r, STATUS_COL, "header", "검수상태", status_cell))
        else:
            ann = ann_map.get(qa)
            want = STATUS_LABELS[ann["s"] if ann else None]
            if status_cell != want:
                diffs.append((r, STATUS_COL, "status", want, status_cell))

    if diffs:
        print(f"FAIL: {len(diffs)} diffs")
        for d in diffs[:20]:
            print(" ", d)
        sys.exit(1)
    longest = max(
        (len(v) for row in orig.iter_rows(values_only=True) for v in row if isinstance(v, str)),
        default=0,
    )
    print(f"OK: diff 0 (rows=501, longest original str cell={longest} chars intact)")


if __name__ == "__main__":
    main()

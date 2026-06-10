# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl"]
# ///
"""data/*.xlsx -> public/data/*.json 변환 + 검증.

실행: uv run scripts/build_data.py
"""

import datetime
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SCHEMA_VERSION = 1
EXPECTED_COLUMNS = [
    "video_idx",
    "QA_idx",
    "영상 유형",
    "영상 제목",
    "영상 링크",
    "질문 유형",
    "질문",
    "객관식 후보1",
    "객관식 후보2",
    "객관식 후보3",
    "객관식 후보4",
    "객관식 후보5",
    "정답",
    "refinement",
    "근거",
    "비고",
]
EXPECTED_ROW_COUNT = 500
MAX_CELL_LEN = 32000  # xlsx 한계 32767 대비 tripwire
QA_IDX_COL = 1
URL_COL = 4

FILES = {
    "drama": ("드라마", "sc04_results_drama.xlsx"),
    "cooking": ("요리", "sc04_results_cooking.xlsx"),
    "assembly": ("조립", "sc04_results_assembly.xlsx"),
}

# src/lib/sc04.ts의 parseYoutubeId와 동일한 패턴 유지
YOUTUBE_PATTERNS = [
    re.compile(r"youtube\.com/shorts/([A-Za-z0-9_-]{11})"),
    re.compile(r"youtube\.com/watch\?.*v=([A-Za-z0-9_-]{11})"),
    re.compile(r"youtu\.be/([A-Za-z0-9_-]{11})"),
    re.compile(r"youtube\.com/embed/([A-Za-z0-9_-]{11})"),
]


def fail(msg: str) -> None:
    print(f"ERROR: {msg}", file=sys.stderr)
    sys.exit(1)


def convert(key: str, label: str, source_file: str, built_at: str) -> dict:
    path = ROOT / "data" / source_file
    if not path.exists():
        fail(f"{path} 없음")
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = [c for c in next(rows_iter)]
    if header != EXPECTED_COLUMNS:
        fail(f"{source_file}: 헤더 불일치\n  expected={EXPECTED_COLUMNS}\n  actual={header}")

    rows = []
    qa_indices = set()
    video_ids = set()
    max_len = 0
    url_warnings = 0
    for raw in rows_iter:
        row = list(raw[: len(EXPECTED_COLUMNS)])
        row = [("" if v is None else v) for v in row]
        for v in row:
            if isinstance(v, str):
                max_len = max(max_len, len(v))
                if len(v) >= MAX_CELL_LEN:
                    fail(f"{source_file}: 셀 길이 {len(v)} >= {MAX_CELL_LEN}")
        qa_idx = row[QA_IDX_COL]
        if qa_idx == "":
            fail(f"{source_file}: QA_idx 빈 값 (row {len(rows) + 2})")
        if qa_idx in qa_indices:
            fail(f"{source_file}: QA_idx 중복 {qa_idx}")
        qa_indices.add(qa_idx)
        url = str(row[URL_COL])
        for pat in YOUTUBE_PATTERNS:
            m = pat.search(url)
            if m:
                video_ids.add(m.group(1))
                break
        else:
            url_warnings += 1
            print(f"WARN: {source_file} QA_idx={qa_idx}: YouTube id 파싱 실패: {url}")
        rows.append(row)

    if len(rows) != EXPECTED_ROW_COUNT:
        fail(f"{source_file}: 행 수 {len(rows)} != {EXPECTED_ROW_COUNT}")

    out = {
        "meta": {
            "schemaVersion": SCHEMA_VERSION,
            "builtAt": built_at,
            "columns": EXPECTED_COLUMNS,
        },
        "key": key,
        "label": label,
        "sourceFile": source_file,
        "rows": rows,
    }
    out_path = ROOT / "public" / "data" / f"{key}.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    size_kb = out_path.stat().st_size / 1024
    print(
        f"{key:9s} rows={len(rows)} videos={len(video_ids)} max_cell_len={max_len} "
        f"url_warn={url_warnings} -> {out_path.relative_to(ROOT)} ({size_kb:.0f} KB)"
    )
    return out


def main() -> None:
    built_at = datetime.datetime.now(datetime.timezone.utc).astimezone().isoformat(timespec="seconds")
    for key, (label, source_file) in FILES.items():
        convert(key, label, source_file, built_at)
    print("OK")


if __name__ == "__main__":
    main()
